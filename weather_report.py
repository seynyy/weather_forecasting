import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

start_year = 0
end_year = 22

count = 21*28

l = {
        0:'B',
        1:'C',
        2:'D',
        3:'E',
        4:'F',
        5:'G',
        6:'H',
        7:'I',
        8:'J',
        9:'K',
        10:'L',
        11:'M',
        12:'N',
        13:'O',
        14:'P'
}

m ={
        'N':0,
        'NNE':22.5,
        'NE':45,
        'ENE':67.5,
        'E':90,
        'ESE':112.5,
        'SE':135,
        'SSE':157.5,
        'S':180,
        'SW':225,   
        'SSW':247.5,
        'W':270,
        'WSW':292.5,
        'WNW':281.5,
        'NW':315,
        'NNW':337.5
}

        

for i in range(0, 22):
    for j in range(1, 29):        
        if i < 10:
                r = requests.get("https://www.ogimet.com/cgi-bin/gsynres?lang=en&ind=12495&ndays=30&ano=200" + str(i)+"&mes=02&day="+ str(j)+"&hora=12&ord=REV&Send=Send","html.parser")                    
        if i >= 10:
                r = requests.get("https://www.ogimet.com/cgi-bin/gsynres?lang=en&ind=12495&ndays=30&ano=20" + str(i)+"&mes=02&day="+ str(j)+"&hora=12&ord=REV&Send=Send","html.parser")

        print(str(round((i*27+j)/count*100, 2)) + "%")

        soup = BeautifulSoup(r.text,'html.parser')

        a = soup.find_all('td', attrs={'align':'right'})
        b = soup.find('td', attrs={'align':'center', 'bgcolor':'#a0a0c0'})

        a.insert(5, b)

        ws['A'+ str(j+3+(i-start_year)*30)] = j

        for k in range(15):
                if a[k] is not None and "504" not in r:              
                        if a[k].string in m:
                                ws[l[k]+str(j+3+(i-start_year)*30)] = float(m[a[k].string])
                        elif a[k].string == '----' or a[k].string == '---':
                                ws[l[k]+str(j+3+(i-start_year)*30)] = 'bd'
                        elif a[k].string == 'Tr':
                                ws[l[k]+str(j+3+(i-start_year)*30)] = 'Tr'
                        elif "504" not in r:
                                ws[l[k]+str(j+3+(i-start_year)*30)] = float(a[k].string)
                        else:
                                for k in range(15):        
                                        ws[l[k]+str(j+3+(i-start_year)*30)] = 'bd'
                                continue

wb.save('weather_report1.xlsx')


input()